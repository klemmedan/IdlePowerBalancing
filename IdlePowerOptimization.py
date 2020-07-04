import matplotlib.pyplot as plt
import numpy
import math
import copy
import openpyxl
from typing import List, Optional

purchase_factor = 1.2


class DataLoader:

    def __init__(self):
        self.wb = openpyxl.load_workbook(filename = "C:\\Users\\Dan\\Documents\\idle_power_balancing\\IdlePowerBalancing.xlsx")
        self.prod_sheet_name = 'prod_base_values'
        self.prod_unlock_sheet_name = 'prod_single_unlocks'
        self.prod_upgrade_sheet_name = 'prod_single_upgrades'

        self.demand_sheet_name = 'demand_base_values'
        self.demand_unlock_sheet_name = 'demand_single_unlocks'
        self.demand_upgrade_sheet_name = 'demand_single_upgrades'

    def load_optimizer(self):
        production_resources = self.load_prod_resources()
        demand_resources = self.load_demand_resources()

    def load_prod_resources(self):
        ws = self.wb[self.prod_sheet_name]
        index = 0
        resources = list()
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=4):
            if row[0].value is None:
                break
            base_increment = row[0].value
            base_cost = row[1].value
            growth_rate = row[2].value
            thresholds, factors = self.load_prod_unlocks(index)
            resource = Resource()
            if index == 0:
                resource.amount = 1
            resource.base_increment = base_increment
            resource.base_cost = base_cost
            resource.growth_rate = growth_rate
            resource.unlock_thresholds = thresholds
            resource.unlock_factors = factors
            resources.append(resource)
            index += 1
        return resources

    def load_prod_unlocks(self, index):
        thresholds = list()
        factors = list()
        ws = self.wb[self.prod_unlock_sheet_name]
        for row in ws.iter_rows(min_row=3):
            threshold = row[0].value
            factor = row[index + 1].value
            thresholds.append(threshold)
            factors.append(factor)
            print(threshold, factor)
        return thresholds, factors

    def load_demand_resources(self):
        ws = self.wb[self.demand_sheet_name]
        index = 0
        resources = list()
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=3):
            if row[0].value is None:
                break
            base_increment = row[0].value
            base_cost = row[1].value
            growth_rate = row[2].value
            thresholds, factors = self.load_demand_unlocks(index)
            resource = Resource()
            if index == 0:
                resource.amount = 1
            resource.base_increment = base_increment
            resource.base_cost = base_cost
            resource.growth_rate = growth_rate
            resource.unlock_thresholds = thresholds
            resource.unlock_factors = factors
            resources.append(resource)
            index += 1
        return resources

    def load_demand_unlocks(self, index):
        thresholds = list()
        factors = list()
        ws = self.wb[self.prod_unlock_sheet_name]
        for row in ws.iter_rows(min_row=3):
            threshold = row[0].value
            factor = row[index + 1].value
            thresholds.append(threshold)
            factors.append(factor)
            print(threshold, factor)
        return thresholds, factors


class Optimizer:

    def __init__(self):
        self.time = list()  # type: List[float]
        self.demand = list()  # type: List[float]
        self.production = list()  # type: List[float]
        self.powerValue = list()  # type: List[float]
        self.income = list()  # type: List[float]
        self.cumulativeProduction = list()  # type: List[float]
        self.cumulativeIncome = list()  # type: List[float]

        self.productionResources = list()  # type: List[Resource]
        self.demandResources = list()  # type: List[Resource]
        self.prodMultiUnlockFactors = list()  # type: List[float]
        self.prodMultiUnlockThresholds = list()  # type: List[float]
        self.demandMultiUnlockFactors = list()  # type: List[float]
        self.demandMultiUnlockThresholds = list()  # type: List[float]
        self.prodMultiUpgrades = UpgradeManager()
        self.demandMultiUpgrades = UpgradeManager()
        self.prodUpgradeManagers = list()  # type: List[UpgradeManager]
        self.demandUpgradeManagers = list()  # type: List[UpgradeManager]

    def run_optimization(self):
        for i in range(10000):
            prod_income = self.total_prod_income()
            demand_income = self.total_demand_income()
            if prod_income > demand_income:
                self.improve_prod()
            else:
                self.improve_demand()

    def improve_prod(self):
        (count_index, count_score) = self.max_prod_resource_score()
        (upgrade_index, upgrade_score) = self.max_prod_upgrade_score()
        if upgrade_score >= count_score:
            self.prodUpgradeManagers[upgrade_index].make_purchase()
        else:
            self.productionResources[count_index].make_purchase()

    def improve_demand(self):
        (count_index, count_score) = self.max_demand_resource_score()
        (upgrade_index, upgrade_score) = self.max_demand_upgrade_score()
        if upgrade_score >= count_score:
            self.demandUpgradeManagers[upgrade_index].make_purchase()
        else:
            self.demandResources[count_index].make_purchase()

    def total_prod_income(self):
        total = 0
        for resource in self.productionResources:
            total = total + resource.increment()
        return total

    def total_demand_income(self):
        total = 0
        for resource in self.productionResources:
            total = total + resource.increment()
        return total

    def max_prod_resource_score(self):
        index = 0
        max_score = 0
        max_index = index
        for resource in self.productionResources:
            cost = resource.cost()
            current_increment_value = resource.increment()
            increment_value_if_purchased = resource.increment_if_purchased()
            score = (increment_value_if_purchased - current_increment_value) / cost

            if score > max_score:
                max_score = score
                max_index = index
            index = index + 1
        return max_index, max_score

    def max_prod_upgrade_score(self):
        index = 0
        max_index = 0
        max_score = 0

        for upgrade_manager in self.prodUpgradeManagers:
            resource_increment = self.productionResources[index].increment()
            cost = upgrade_manager.next_cost()
            current_factor = upgrade_manager.current_factor()
            next_factor = upgrade_manager.next_factor()
            score = resource_increment * (next_factor/current_factor - 1) / cost
            if score > max_score:
                max_score = score
                max_index = index
            index = index + 1
        return max_index, max_score

    def max_demand_resource_score(self):
        index = 0
        max_score = 0
        max_index = index
        for resource in self.demandResources:
            cost = resource.cost()
            current_increment_value = resource.increment()
            increment_value_if_purchased = resource.increment_if_purchased()
            score = (increment_value_if_purchased - current_increment_value) / cost

            if score > max_score:
                max_score = score
                max_index = index
            index = index + 1
        return max_index, max_score

    def max_demand_upgrade_score(self):
        index = 0
        max_index = 0
        max_score = 0

        for upgrade_manager in self.demandUpgradeManagers:
            resource_increment = self.demandResources[index].increment()
            cost = upgrade_manager.next_cost()
            current_factor = upgrade_manager.current_factor()
            next_factor = upgrade_manager.next_factor()
            score = resource_increment * (next_factor/current_factor - 1) / cost
            if score > max_score:
                max_score = score
                max_index = index
            index = index + 1
        return max_index, max_score


class Resource:

    def __init__(self):
        self.amount = 0
        self.base_increment = 1
        self.base_cost = 10
        self.growth_rate = 1.1
        # Threshold to achieve an unlock. Start at 0 (corresponds to 1x factor, and go from there)
        self.unlock_thresholds = list()
        # Factor is cumulative and starts at 1 (i.e. 1, 2, 4, 12, 36, etc.)
        self.unlock_factors = list()

    def increment(self):
        unlock_index = self.unlock_index()
        unlock_factor = self.unlock_factors[unlock_index]
        return self.amount * self.base_increment * unlock_factor

    def cost(self):
        n = self.num_to_purchase()
        total_cost = self.growth_rate * (1 - self.growth_rate ** (n + self.amount)) / (1 - self.growth_rate)
        previous_cost = self.growth_rate * (1 - self.growth_rate ** self.amount) / (1 - self.growth_rate)
        diff_cost = total_cost - previous_cost
        return diff_cost

    def unlock_index(self):
        index = 0
        for threshold in self.unlock_thresholds:
            if self.amount >= threshold:
                index = index + 1
            else:
                break
        return index

    def next_unlock_threshold(self):
        return self.unlock_thresholds[self.unlock_index() + 1]

    def num_to_purchase(self):
        to_next_threshold = self.next_unlock_threshold() - self.amount
        base_increase = math.floor(self.amount*(purchase_factor - 1))
        return min(base_increase, to_next_threshold)

    def increment_if_purchased(self):
        bought_resource = copy.copy(self)
        bought_resource.amount = bought_resource.amount + self.num_to_purchase()
        return bought_resource.increment()

    def make_purchase(self):
        self.amount = self.amount + self.num_to_purchase()


class UpgradeManager:

    def __init__(self):
        self.current_index = 0

        #
        self.upgrade_costs = list()

        # Cumulative factor, start at 1, then multiples all the way up.
        self.upgrade_factors = list()

    def current_cost(self):
        return self.upgrade_costs[self.current_index]

    def current_factor(self):
        return self.upgrade_factors[self.current_index]

    def next_cost(self):
        return self.upgrade_costs[self.current_index + 1]

    def next_factor(self):
        return self.upgrade_factors[self.current_index + 1]

    def make_purchase(self):
        self.current_index += 1



DataLoader().load_prod_resources()

opt = Optimizer()