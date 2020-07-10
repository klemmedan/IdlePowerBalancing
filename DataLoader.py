
import openpyxl
from IdlePowerOptimization import Optimizer, UpgradeManager, Resource, PowerValue
from typing import List, Optional


class DataLoader:

    def __init__(self):
        self.production_resources = list()  # type: List[Resource]
        self.demand_resources = list()  # type: List[Resource]
        self.production_upgrade_managers = list()  # type: List[UpgradeManager]
        self.demand_upgrade_managers = list()  # type: List[UpgradeManager]
        self.production_multi_unlock_factors = list()
        self.production_multi_unlock_thresholds = list()
        self.production_multi_upgrade_manager = UpgradeManager()
        self.demand_multi_unlock_factors = list()
        self.demand_multi_unlock_thresholds = list()
        self.demand_multi_upgrade_manager = UpgradeManager()
        self.power_value = PowerValue()
        self.power_value_upgrade_manager = UpgradeManager()

        self.optimizer = Optimizer()

        self.wb = openpyxl.load_workbook(filename="IdlePowerBalancing.xlsx", data_only=True)
        self.prod_sheet_name = 'prod_base_values'
        self.prod_unlock_sheet_name = 'prod_single_unlocks'
        self.prod_upgrade_sheet_name = 'prod_single_upgrades'
        self.prod_multi_sheet_name = 'prod_multi'

        self.demand_sheet_name = 'demand_base_values'
        self.demand_unlock_sheet_name = 'demand_single_unlocks'
        self.demand_upgrade_sheet_name = 'demand_single_upgrades'
        self.demand_manual_unlocks_sheet_name = 'demand_manual_unlocks'
        self.demand_manual_upgrades_sheet_name = 'demand_manual_upgrades'
        self.demand_multi_sheet_name = 'demand_multi'

        self.power_value_sheet_name = 'power_value'

    def load_optimizer(self):
        # type: () -> Optimizer
        self.production_resources = self.load_prod_resources()
        self.demand_resources = self.load_demand_resources()
        self.production_upgrade_managers = self.load_prod_upgrade_managers()
        self.demand_upgrade_managers = self.load_demand_upgrade_managers()

        self.production_multi_upgrade_manager = self.load_multi_prod_upgrade_manager()
        self.production_multi_unlock_thresholds, self.production_multi_unlock_factors = self.load_multi_prod_unlocks()
        self.demand_multi_upgrade_manager = self.load_multi_demand_upgrade_manager()
        self.demand_multi_unlock_thresholds, self.demand_multi_unlock_factors = self.load_multi_demand_unlocks()

        self.power_value = self.load_power_value()
        self.power_value_upgrade_manager = self.load_power_value_upgrade_manager()

        optimizer = Optimizer()
        optimizer.prod_resources = self.production_resources
        optimizer.demand_resources = self.demand_resources
        optimizer.prod_multi_unlock_factors = self.production_multi_unlock_factors
        optimizer.prod_multi_unlock_thresholds = self.production_multi_unlock_thresholds
        optimizer.prod_multi_upgrades = self.production_multi_upgrade_manager
        optimizer.prod_upgrade_managers = self.production_upgrade_managers
        optimizer.demand_multi_unlock_factors = self.demand_multi_unlock_factors
        optimizer.demand_multi_unlock_thresholds = self.demand_multi_unlock_thresholds
        optimizer.demand_upgrade_managers = self.demand_upgrade_managers
        optimizer.demand_multi_upgrades = self.demand_multi_upgrade_manager
        optimizer.power_value_upgrade_manager = self.power_value_upgrade_manager
        optimizer.power_value = self.power_value

        return optimizer

    def load_prod_resources(self):
        ws = self.wb[self.prod_sheet_name]
        index = 0
        resources = list()
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=4):
            if row[0].value is None:
                break
            base_increment = row[0].value
            base_cost = row[1].value
            growth_rate = row[2].value
            thresholds, factors = self.load_prod_unlocks(index)
            resource = Resource()
            if index == 0:
                resource.amount = 1
            resource.base_increment = base_increment
            resource.base_cost = base_cost
            resource.growth_rate = growth_rate
            resource.unlock_thresholds = thresholds
            resource.unlock_factors = factors
            resource.initialize()
            resources.append(resource)
            index += 1
        return resources

    def load_prod_unlocks(self, index):
        ws = self.wb[self.prod_unlock_sheet_name]
        thresholds = [0] + [row[0].value for row in ws.iter_rows(min_row=3)]
        base_factors = [row[index+1].value for row in ws.iter_rows(min_row=3)]
        factors = self.cumulative_factors(base_factors)
        return thresholds, factors

    def load_demand_resources(self):
        ws = self.wb[self.demand_sheet_name]
        index = 0
        resources = list()
        for row in ws.iter_rows(min_row=4, min_col=2):
            if row[2].value is None:
                break
            base_increment = row[0].value
            base_cost = row[1].value
            growth_rate = row[2].value
            thresholds, factors = self.load_demand_unlocks(index)
            resource = Resource()
            if index == 0:
                resource.amount = 1
            resource.base_increment = base_increment
            resource.base_cost = base_cost
            resource.growth_rate = growth_rate
            resource.unlock_thresholds = thresholds
            resource.unlock_factors = factors
            resource.initialize()
            resources.append(resource)
            index += 1
        return resources

    def load_demand_unlocks(self, index):
        if index == 0:
            return self.load_manual_unlocks()
        else:
            ws = self.wb[self.demand_unlock_sheet_name]
            thresholds = [0] + [row[0].value for row in ws.iter_rows(min_row=4)]
            base_factors = [row[index].value for row in ws.iter_rows(min_row=4)]
            factors = self.cumulative_factors(base_factors)
            return thresholds, factors

    def load_manual_unlocks(self):
        ws = self.wb[self.demand_manual_unlocks_sheet_name]
        thresholds = [0] + [row[0].value for row in ws.iter_rows(min_row=4)]
        base_factors = [row[1].value for row in ws.iter_rows(min_row=4)]
        factors = self.cumulative_factors(base_factors)
        return thresholds, factors

    def load_prod_upgrade_managers(self):
        num_resources = len(self.production_resources)
        return [self.load_prod_upgrade_manager(i) for i in range(num_resources)]

    def load_prod_upgrade_manager(self, index):
        ws = self.wb[self.prod_upgrade_sheet_name]
        costs = [0] + [row[index].value for row in ws.iter_rows(min_row=3, min_col=2)]
        base_factors = [row[index+int(len(row)/2)].value for row in ws.iter_rows(min_row=3, min_col=2)]
        factors = self.cumulative_factors(base_factors)

        return self.make_upgrade_manager(costs, factors)

    def load_demand_upgrade_managers(self):
        num_resources = len(self.demand_resources)
        return [self.load_demand_upgrade_manager(i) for i in range(num_resources)]

    def load_demand_upgrade_manager(self, index):
        if index == 0:
            return self.load_manual_upgrade_manager()
        else:
            ws = self.wb[self.demand_upgrade_sheet_name]
            costs = [0] + [row[index-1].value for row in ws.iter_rows(min_row=4, min_col=2)]
            base_factors = [row[index+int(len(row)/2-1)].value for row in ws.iter_rows(min_row=4, min_col=2)]
            factors = self.cumulative_factors(base_factors)

            return self.make_upgrade_manager(costs, factors)

    def load_manual_upgrade_manager(self):
        ws = self.wb[self.demand_manual_upgrades_sheet_name]
        costs = [0] + [row[0].value for row in ws.iter_rows(min_row=4)]
        base_factors = [row[1].value for row in ws.iter_rows(min_row=4)]
        factors = self.cumulative_factors(base_factors)
        return self.make_upgrade_manager(costs, factors)

    def load_multi_prod_unlocks(self):
        ws = self.wb[self.prod_multi_sheet_name]
        thresholds = [0] + [row[0].value for row in ws.iter_rows(min_row=3) if row[0].value is not None]
        base_factors = [row[1].value for row in ws.iter_rows(min_row=3) if row[0].value is not None]
        factors = self.cumulative_factors(base_factors)
        return thresholds, factors

    def load_multi_prod_upgrade_manager(self):
        ws = self.wb[self.prod_multi_sheet_name]
        costs = [0] + [row[3].value for row in ws.iter_rows(min_row=3)]
        base_factors = [row[4].value for row in ws.iter_rows(min_row=3)]
        factors = self.cumulative_factors(base_factors)
        return self.make_upgrade_manager(costs, factors)

    def load_multi_demand_unlocks(self):
        ws = self.wb[self.demand_multi_sheet_name]
        thresholds = [0] + [row[0].value for row in ws.iter_rows(min_row=4) if row[0].value is not None]
        base_factors = [row[1].value for row in ws.iter_rows(min_row=4)if row[0].value is not None]
        factors = self.cumulative_factors(base_factors)
        return thresholds, factors

    def load_multi_demand_upgrade_manager(self):
        ws = self.wb[self.demand_multi_sheet_name]
        costs = [0] + [row[2].value for row in ws.iter_rows(min_row=4)]
        base_factors = [row[3].value for row in ws.iter_rows(min_row=4)]
        factors = self.cumulative_factors(base_factors)
        return self.make_upgrade_manager(costs, factors)

    def load_power_value(self):
        ws = self.wb[self.power_value_sheet_name]
        power_value = PowerValue()
        power_value.amount_growth_rate = ws["C5"].value
        power_value.amount_second_growth_rate = ws["D5"].value
        power_value.base_cost = ws["E5"].value
        power_value.cost_growth_rate = ws["F5"].value
        return power_value

    def load_power_value_upgrade_manager(self):
        ws = self.wb[self.power_value_sheet_name]
        costs = [0] + [row[0].value for row in ws.iter_rows(min_row=4)]
        base_factors = [row[1].value for row in ws.iter_rows(min_row=4)]
        factors = self.cumulative_factors(base_factors)
        return self.make_upgrade_manager(costs, factors)

    @staticmethod
    def cumulative_factors(base_factors):
        factors = [1]
        for f in base_factors:
            factors.append(factors[-1]*f)
        return factors

    @staticmethod
    def make_upgrade_manager(costs, factors):
        manager = UpgradeManager()
        manager.upgrade_costs = costs
        manager.upgrade_factors = factors
        return manager

